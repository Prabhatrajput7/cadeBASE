import openpyxl

class HomeRelated:
    text_to_pass = [{"fname":"RKO","ename":"gmail@gmail","gender":"Male"},{"fname":"AA","ename":"gmail@gmail","gender":"Female"}]

    @staticmethod
    def get_testdataEX(test_case_name):
        dict ={}
        where = openpyxl.load_workbook("C:\\Users\\Prabh\\Documents\\pdemo.xlsx")
        sheet = where.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

                return [dict]

    def get_testdataEXC(self, test_case_name):
        dict ={}
        where = openpyxl.load_workbook("C:\\Users\\Prabh\\Documents\\pdemo.xlsx")
        sheet = where.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                return [dict]

apple = HomeRelated()
# apple.get_testdataEXC('Testcase1')

