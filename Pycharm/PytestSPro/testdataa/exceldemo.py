import openpyxl
where = openpyxl.load_workbook("C:\\Users\\Prabh\\Documents\\pdemo.xlsx")
sheet = where.active
dict ={}
cell = sheet.cell(row=1,column=2)
# print(cell.value)
# sheet.cell(row=2,column=2).value = 'Don'
# print(sheet.cell(row=2,column=2).value)
# print(sheet.max_row)
#print(sheet['A1'].value)

# for i in range(1,sheet.max_row+1):
#     if sheet.cell(row=i,column=1).value == 'Testcase1':
#         for j in range(1,sheet.max_column+1):
#             print(sheet.cell(row=i,column=j).value)

for i in range(1,sheet.max_row+1):
    for j in range(2,sheet.max_column+1):
        dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
    print(dict)







